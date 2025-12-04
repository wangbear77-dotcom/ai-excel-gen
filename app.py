import streamlit as st
import google.generativeai as genai
from google.generativeai.types import HarmCategory, HarmBlockThreshold
import pandas as pd
import io
import random
import datetime
import traceback
import re

# 預先檢查環境
try:
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    st.error("請先安裝 openpyxl: pip install openpyxl")

# --- 1. 頁面設定 ---
st.set_page_config(page_title="Excel 菜鳥救星", page_icon="🐣", layout="centered")

if 'user_prompt' not in st.session_state:
    st.session_state['user_prompt'] = ''

# --- 2. 標題 ---
st.title("🐣 Excel 菜鳥救星")
st.markdown("不懂公式？不會排版？讓 AI 幫你寫一個 **「還不錯的草稿」**！")

st.warning("""
**⚠️ 使用前請詳閱：**
* 這是輔助工具，適合生成 **「基礎到中等難度」** 的表格。
* **太專業的需求** (如複雜連動、VBA) AI 可能會暈倒。
* **可能有小 Bug**：建議下載後 **「人工檢查」** 公式。
* **如果不滿意**：請調整描述再試一次 (AI 每次發揮都不太一樣)。
""")

# --- 3. 側邊欄：設定 (V6.2 的隨機 Key 邏輯完整保留) ---
with st.sidebar:
    st.header("⚙️ 設定與權限")
    
    current_api_key = None
    
    # 嘗試讀取後台 Secrets
    try:
        if "API_KEYS" in st.secrets:
            keys = st.secrets["API_KEYS"]
            if isinstance(keys, list) and len(keys) > 0:
                current_api_key = random.choice(keys)
        elif "GEMINI_API_KEY" in st.secrets:
            current_api_key = st.secrets["GEMINI_API_KEY"]
    except FileNotFoundError:
        pass 
    except Exception:
        pass 

    # 如果後台沒抓到 Key，顯示手動輸入框
    if not current_api_key:
        current_api_key = st.text_input("開發者專用 Key (用戶看不到)", type="password")
        if not current_api_key:
            st.warning("⚠️ 系統維護中 (未設定後台金鑰)")

    st.divider()

    # 打賞區
    st.subheader("☕ 鼓勵開發者")
    st.markdown("如果這個工具幫你省了時間，歡迎請我喝杯咖啡！")
    
    st.markdown(
        """
        <a href="https://www.buymeacoffee.com/wangbear77" target="_blank">
            <img src="https://cdn.buymeacoffee.com/buttons/v2/default-yellow.png" alt="Buy Me A Coffee" style="height: 40px !important;width: 150px !important;" >
        </a>
        """,
        unsafe_allow_html=True
    )

    st.divider()
    
    # 懶人樣板
    st.write("⚡ **快速樣板 (點擊填寫)：**")
    if st.button("💰 個人記帳表"): st.session_state['user_prompt'] = "幫我做一個2025年個人記帳表。欄位：日期、類別、項目、金額、付款方式。請生成10筆範例。公式要求：計算本月總支出、分類小計。美化：標題深藍底白字，金額加$符號。"
    if st.button("📦 商品庫存表"): st.session_state['user_prompt'] = "幫我做一個庫存管理表。欄位：商品編號、名稱、進貨價、售價、庫存量、庫存總值(公式：進貨價*庫存量)。請生成10筆範例。美化：標題深綠底，金額加千分位。"
    if st.button("🛒 網拍訂單表"): st.session_state['user_prompt'] = "幫我做一個電商訂單管理表。欄位：訂單編號、平台(蝦皮/官網)、商品、單價、數量、手續費(蝦皮8%/官網2%)、實收金額。請生成10筆。公式要求：用IF判斷手續費，退貨實收為0。美化：標題亮橘底。"

    model_choice = st.selectbox("模型選擇", ["gemini-2.5-flash", "gemini-2.5-pro"])

# --- 4. 核心邏輯 (加入 V6.4 公式強制鎖定) ---
def sanitize_code(code):
    lines = code.split('\n')
    cleaned_lines = []
    for line in lines:
        if "openpyxl.worksheet.conditional_formatting" in line: continue
        if "openpyxl.formatting.rule" in line: continue
        if "FormulaRule" in line: continue
        cleaned_lines.append(line)
    return '\n'.join(cleaned_lines)

def generate_and_fix_code(user_prompt, key, model_name):
    try:
        genai.configure(api_key=key)
        safety_settings = {HarmCategory.HARM_CATEGORY_HARASSMENT: HarmBlockThreshold.BLOCK_NONE, HarmCategory.HARM_CATEGORY_HATE_SPEECH: HarmBlockThreshold.BLOCK_NONE, HarmCategory.HARM_CATEGORY_SEXUALLY_EXPLICIT: HarmBlockThreshold.BLOCK_NONE, HarmCategory.HARM_CATEGORY_DANGEROUS_CONTENT: HarmBlockThreshold.BLOCK_NONE}
        model = genai.GenerativeModel(model_name, generation_config=genai.types.GenerationConfig(max_output_tokens=8000, temperature=0.0)) 
        
        # 🔥 V6.4 修正：強制 AI 只輸出公式字串，不准先計算結果
        base_prompt = f"""
        你是一位 Python Excel 自動化專家。需求："{user_prompt}"
        請寫一段 **完整且可執行** 的 Python 代碼。
        
        【嚴格代碼規範】：
        1. **Imports**：務必包含 io, random, datetime, pandas, openpyxl 相關模組。
        2. **核心邏輯**：建立 wb = Workbook()，填入資料與公式，美化樣式。
        
        3. **公式寫法 (CRITICAL)**：
           - **絕對禁止** 在 Python 代碼中算出結果 (例如: amount = 500 * 1.05)。
           - 所有的 Excel 公式字串 **必須以等於號 (=) 開頭**。
           - 請使用 `ws.cell(row=i, column=j, value=f'=...')` 的方式寫入公式。
           - **最終檢查**：在儲存前，檢查所有儲存格，如果公式內容是像 SUM, IF, PMT 開頭但沒有等號，請自動補上。
        
        4. **禁止模組**：不要使用 openpyxl.formatting 或 conditional_formatting (請用迴圈變色)。
        5. **關鍵步驟**：最後請將檔案儲存到變數 `output_buffer = io.BytesIO()`，並 `wb.save(output_buffer)`。
        6. **禁止事項**：只輸出 Python 代碼，不要 markdown。
        """
        
        current_prompt = base_prompt
        max_retries = 3 
        
        for attempt in range(max_retries):
            response = model.generate_content(current_prompt, safety_settings=safety_settings)
            
            if not response.parts: return None, f"AI 拒絕生成 (Finish Reason: {response.candidates[0].finish_reason})。"
                
            raw_code = response.text
            clean_code = raw_code.replace("```python", "").replace("```", "").strip()
            if not clean_code.startswith("import") and not clean_code.startswith("from"):
                 import_pos = clean_code.find("import")
                 if import_pos != -1: clean_code = clean_code[import_pos:]
            
            clean_code = sanitize_code(clean_code)

            try:
                test_vars = {}
                exec(clean_code, globals(), test_vars)
                if 'output_buffer' in test_vars: return clean_code, None
                else: raise Exception("代碼執行成功但未產生 output_buffer 變數")
            except Exception as e:
                error_msg = str(e)
                print(f"Retry {attempt+1}: {error_msg}")
                # 錯誤回報：特別強調要寫 =
                current_prompt += f"\n\n【系統回報】：程式碼執行失敗，錯誤訊息：{error_msg}。\n請修正代碼，**務必確保公式字串以等於號 (=) 開頭**，然後重新輸出。"
                
        return None, "AI 嘗試修復了 3 次但仍然失敗，請嘗試簡化您的需求。"
    except Exception as e:
        return None, str(e)

# --- 5. 主介面 ---
with st.expander("💡 怎麼樣才能做出完美的表格？ (點我看秘訣)"):
    st.markdown("""
    **黃金許願公式：**
    > **我要做什麼表 + 有哪些欄位 + 資料邏輯/公式 + 美化要求**
    
    **❌ 壞範例：**
    "幫我做一個記帳表。" (AI 不知道你要記什麼，可能會做得很簡陋)
    
    **✅ 好範例：**
    "幫我做一個**家庭收支表**。
    欄位要有：**日期、項目、金額、類別**。
    請幫我造 **10 筆** 隨機資料。
    最下面要用公式幫我算**總金額**。
    標題請用**深綠色底**，金額要有**錢字號**。"
    """)

user_input = st.text_area("請輸入需求 (或點擊左側快速樣板)：", value=st.session_state['user_prompt'], height=150, placeholder="例如：幫我做一個房東收租表...")

if st.button("✨ 生成專業表格", type="primary"):
    if not current_api_key:
        st.error("⚠️ 系統維護中 (未設定後台金鑰)")
    elif not user_input:
        st.warning("⚠️ 請輸入需求")
    else:
        spinner_text = f"🤖 AI 正在製作中 (公式強力校正模式)..."
        with st.spinner(spinner_text):
            code, error_msg = generate_and_fix_code(user_input, current_api_key, model_choice)
            if code:
                try:
                    local_vars = {}
                    exec(code, globals(), local_vars)
                    
                    if 'output_buffer' in local_vars:
                        excel_data = local_vars['output_buffer']
                        file_name = f"excel_gen_{datetime.datetime.now().strftime('%H%M%S')}.xlsx"
                        st.download_button(label="📥 下載 Excel (.xlsx)", data=excel_data, file_name=file_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                        st.success("🎉 完成！請檢查 Excel 內容。")
                    else: st.error("生成失敗。")
                except Exception as e:
                    st.error(f"錯誤：{e}")
                    with st.expander("查看代碼"): st.code(code, language='python')
            else:
                st.error("連線或修復失敗，請稍後再試。")
                if error_msg: st.error(error_msg)

# --- 6. 頁尾 ---
st.divider()
st.caption("Excel Rookie Savior V6.4 (Final Formula Lock)")
